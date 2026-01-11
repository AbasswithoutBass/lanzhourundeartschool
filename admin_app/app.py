#!/usr/bin/env python3
import os
import json
import datetime
import shutil
import random
import re
import io
import html as _html
import urllib.request
import urllib.parse
from functools import wraps
from pathlib import Path
import importlib.util
from typing import Iterable

try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:  # pragma: no cover
    Image = ImageDraw = ImageFont = None

try:
    import qrcode
except Exception:  # pragma: no cover
    qrcode = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file


ROOT = Path(__file__).resolve().parents[1]
DATA_TEACHERS = ROOT / 'data' / 'teachers.json'
DATA_STUDENTS = ROOT / 'data' / 'students.json'
DATA_PORTAL = ROOT / 'data' / 'portal_posts.json'

_TEACHERS_MANAGE = None
_STUDENTS_MANAGE = None


def _load_teachers_manage_module():
    """动态加载 modules/teachers/manage.py 以复用其清洗/去重/归类规则。

    该目录不是 Python package（没有 __init__.py），所以用 importlib 载入。
    """
    global _TEACHERS_MANAGE
    if _TEACHERS_MANAGE is not None:
        return _TEACHERS_MANAGE

    mod_path = ROOT / 'modules' / 'teachers' / 'manage.py'
    spec = importlib.util.spec_from_file_location('teachers_manage', str(mod_path))
    if not spec or not spec.loader:
        raise RuntimeError('无法加载 teachers/manage.py')
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    _TEACHERS_MANAGE = mod
    return _TEACHERS_MANAGE


def _load_students_manage_module():
    """动态加载 modules/students/manage.py 以复用校验/ID 生成/字段规范化逻辑。"""
    global _STUDENTS_MANAGE
    if _STUDENTS_MANAGE is not None:
        return _STUDENTS_MANAGE

    mod_path = ROOT / 'modules' / 'students' / 'manage.py'
    spec = importlib.util.spec_from_file_location('students_manage', str(mod_path))
    if not spec or not spec.loader:
        raise RuntimeError('无法加载 students/manage.py')
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    _STUDENTS_MANAGE = mod
    return _STUDENTS_MANAGE


def _safe_int(v, default=10**9):
    try:
        return int(v)
    except Exception:
        return default


def create_app():
    app = Flask(__name__, template_folder='templates', static_folder='static')

    # 管理后台属于本地开发/内网工具：默认开启模板自动刷新，避免改了 HTML/CSS 但页面还是旧的。
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.jinja_env.auto_reload = True
    # 尽量避免浏览器/代理缓存（尤其是管理后台迭代 UI 时）
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

    # 你必须设置一个随机的 SECRET_KEY（用于 session cookie）
    app.secret_key = os.environ.get('ADMIN_SECRET_KEY') or 'CHANGE_ME_ADMIN_SECRET_KEY'

    def is_authed() -> bool:
        return bool(session.get('admin_authed'))

    def login_required(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not is_authed():
                return redirect(url_for('login', next=request.path))
            return fn(*args, **kwargs)
        return wrapper

    def backup_file(path: Path):
        if not path.exists():
            return
        ts = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        bak = path.with_suffix(path.suffix + '.bak.' + ts)
        shutil.copy2(path, bak)

    def load_json(path: Path):
        if not path.exists():
            return []
        return json.loads(path.read_text(encoding='utf-8'))

    def write_json(path: Path, data):
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_file(path)
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

    def _now_iso() -> str:
        return datetime.datetime.now().isoformat(timespec='seconds')

    def _make_id(prefix: str = 'p') -> str:
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        rand = random.randint(1000, 9999)
        return f'{prefix}_{ts}_{rand}'

    def _norm_tags(s: str) -> list[str]:
        s = str(s or '').replace('，', ',')
        out: list[str] = []
        for part in s.split(','):
            t = str(part or '').strip()
            if t:
                out.append(t)
        # de-dup while keeping order
        seen: set[str] = set()
        uniq: list[str] = []
        for t in out:
            if t in seen:
                continue
            uniq.append(t)
            seen.add(t)
        return uniq

    def _load_portal_posts() -> list[dict]:
        data = load_json(DATA_PORTAL)
        return data if isinstance(data, list) else []

    def _save_portal_posts(posts: list[dict]):
        write_json(DATA_PORTAL, posts)

    def _find_post(posts: list[dict], pid: str) -> dict | None:
        pid = str(pid or '').strip()
        for p in posts:
            if str(p.get('id') or '') == pid:
                return p
        return None

    def _safe_filename(name: str) -> str:
        name = str(name or '').strip().replace(' ', '_')
        name = re.sub(r'[^A-Za-z0-9_\-\.]+', '', name)
        name = re.sub(r'\.+', '.', name)
        return name[:120] if name else 'image'

    def _html_to_text_simple(html_s: str) -> str:
        """A tiny HTML -> text helper for poster generation.

        Keeps basic block breaks so the poster has readable paragraphs.
        """
        s = str(html_s or '')
        # normalize common breaks
        s = re.sub(r'(?is)<\s*br\s*/?\s*>', '\n', s)
        s = re.sub(r'(?is)</\s*p\s*>', '\n\n', s)
        s = re.sub(r'(?is)</\s*div\s*>', '\n', s)
        s = re.sub(r'(?is)</\s*li\s*>', '\n', s)
        # turn list items into bullets
        s = re.sub(r'(?is)<\s*li\b[^>]*>', '• ', s)
        # strip the rest tags
        s = re.sub(r'(?is)<[^>]+>', '', s)
        s = _html.unescape(s)
        # cleanup
        s = s.replace('\r\n', '\n').replace('\r', '\n')
        s = re.sub(r'\n{3,}', '\n\n', s)
        return s.strip()

    def _pick_cjk_font(size: int):
        if ImageFont is None:
            return None

        candidates = [
            # macOS
            '/System/Library/Fonts/PingFang.ttc',
            '/System/Library/Fonts/STHeiti Medium.ttc',
            '/System/Library/Fonts/Hiragino Sans GB.ttc',
            # common linux paths
            '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
            '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
            '/usr/share/fonts/truetype/noto/NotoSansCJKsc-Regular.otf',
        ]
        for p in candidates:
            try:
                if Path(p).exists():
                    return ImageFont.truetype(p, size=size)
            except Exception:
                continue

        try:
            return ImageFont.load_default()
        except Exception:
            return None

    def _wrap_text(draw, text: str, font, max_width: int) -> list[str]:
        """Wrap text to fit in pixel width."""
        if not text:
            return []
        lines: list[str] = []
        for para in str(text).split('\n'):
            para = para.rstrip()
            if not para:
                lines.append('')
                continue

            cur = ''
            for ch in para:
                test = cur + ch
                w = draw.textlength(test, font=font) if hasattr(draw, 'textlength') else draw.textbbox((0, 0), test, font=font)[2]
                if w <= max_width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                        cur = ch
                    else:
                        # single char too wide (unlikely), force
                        lines.append(ch)
                        cur = ''
            if cur:
                lines.append(cur)
        return lines

    def _load_image_any(src: str):
        if Image is None:
            return None
        src = str(src or '').strip()
        if not src:
            return None

        # local project-relative path
        try:
            rp = _norm_json_path(src)
            p = ROOT / rp
            if p.exists() and p.is_file():
                return Image.open(p).convert('RGB')
        except Exception:
            pass

        # remote url
        if src.startswith('http://') or src.startswith('https://'):
            try:
                req = urllib.request.Request(src, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=6) as resp:
                    data = resp.read()
                return Image.open(io.BytesIO(data)).convert('RGB')
            except Exception:
                return None

        return None

    def _make_qr(url: str, size: int = 320):
        if Image is None or qrcode is None:
            return None
        url = str(url or '').strip()
        if not url:
            return None
        try:
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=2,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img = img.convert('RGB')
            img = img.resize((size, size))
            return img
        except Exception:
            return None

    def _clean_upload_relpath(p: str) -> str:
        p = (p or '').replace('\\', '/').lstrip('/')
        p = re.sub(r'/+', '/', p)
        parts = [x for x in p.split('/') if x not in ('', '.')]
        if any(x == '..' for x in parts):
            raise ValueError('path contains ..')
        return '/'.join(parts)

    def _normalize_webkitdirectory_upload(files: Iterable) -> dict:
        """Build a {relpath -> FileStorage} map.

        Browsers usually prefix all filenames with the selected folder name.
        We strip the common first path component when possible.
        """
        cleaned = []
        for f in files:
            if not getattr(f, 'filename', None):
                continue
            cleaned.append(_clean_upload_relpath(f.filename))

        if not cleaned:
            return {'prefix': '', 'files': {}}

        first_parts = [p.split('/')[0] for p in cleaned if '/' in p]
        prefix = ''
        if first_parts and len(set(first_parts)) == 1:
            candidate = first_parts[0]
            # only strip if all paths either start with candidate/ or equal candidate
            if all(p == candidate or p.startswith(candidate + '/') for p in cleaned):
                prefix = candidate

        out: dict[str, object] = {}
        for f in files:
            if not getattr(f, 'filename', None):
                continue
            p = _clean_upload_relpath(f.filename)
            if prefix and (p == prefix or p.startswith(prefix + '/')):
                p = p[len(prefix) + 1:] if p != prefix else ''
            if not p:
                continue
            out[p] = f

        return {'prefix': prefix, 'files': out}

    def _read_json_from_upload(file_map: dict, relpath: str):
        f = file_map.get(relpath)
        if not f:
            raise FileNotFoundError(relpath)
        try:
            f.stream.seek(0)
        except Exception:
            pass
        raw = f.read()
        try:
            txt = raw.decode('utf-8')
        except Exception:
            txt = raw.decode('utf-8-sig', errors='replace')
        return json.loads(txt)

    def _save_uploaded_file(file_map: dict, src_relpath: str, dest: Path):
        f = file_map.get(src_relpath)
        if not f:
            raise FileNotFoundError(src_relpath)
        dest.parent.mkdir(parents=True, exist_ok=True)
        try:
            f.stream.seek(0)
        except Exception:
            pass
        with dest.open('wb') as w:
            shutil.copyfileobj(f.stream, w)

    def _find_upload_by_basename(file_map: dict, filename: str) -> str | None:
        """Return a unique key whose basename matches filename.

        If multiple matches exist, return None and let caller raise an error.
        """
        filename = str(filename or '').strip()
        if not filename:
            return None
        matches = [k for k in file_map.keys() if Path(k).name == filename]
        if len(matches) == 1:
            return matches[0]
        return None

    def _is_allowed_image(path: str) -> bool:
        ext = (Path(path).suffix or '').lower()
        return ext in {'.jpg', '.jpeg', '.png', '.webp'}

    def _norm_json_path(p: str) -> str:
        p = str(p or '').strip().replace('\\', '/')
        p = re.sub(r'/+', '/', p)
        if p.startswith('/'):
            raise ValueError('absolute path not allowed')
        if '..' in p.split('/'):
            raise ValueError('path contains ..')
        return p

    def _merge_by_id(existing: list[dict], incoming: list[dict], *, mode: str) -> tuple[list[dict], int, int]:
        idx = {str(x.get('id')): x for x in existing if x.get('id')}
        created = 0
        updated = 0
        for item in incoming:
            iid = str(item.get('id') or '').strip()
            if not iid:
                continue
            if iid not in idx:
                existing.append(item)
                idx[iid] = item
                created += 1
                continue

            if mode == 'replace':
                # replace in-place to keep list order stable
                cur = idx[iid]
                cur.clear()
                cur.update(item)
            else:
                # merge: update keys present in incoming
                cur = idx[iid]
                for k, v in item.items():
                    if k == 'id':
                        continue
                    cur[k] = v
            updated += 1
        return existing, created, updated

    def _read_xlsx_rows(file_storage, *, required_cols: list[str]) -> list[dict]:
        if load_workbook is None:
            raise RuntimeError('缺少 openpyxl，无法解析 .xlsx（请先安装 openpyxl）')
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
        wb = load_workbook(file_storage.stream, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(x).strip() if x is not None else '' for x in rows[0]]
        header_map = {h: idx for idx, h in enumerate(header) if h}

        missing = [c for c in required_cols if c not in header_map]
        if missing:
            raise ValueError('缺少列：' + ', '.join(missing))

        out: list[dict] = []
        for r in rows[1:]:
            if r is None:
                continue
            d: dict = {}
            empty = True
            for h, idx in header_map.items():
                v = r[idx] if idx < len(r) else None
                if v is None:
                    continue
                if isinstance(v, str):
                    v = v.strip()
                    if v == '':
                        continue
                d[h] = v
                empty = False
            if not empty:
                out.append(d)
        return out

    def _boolish(v) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in {'1', 'true', 'yes', 'y', '是', '对', '√', '已', 'ok'}

    def _enforce_mgmt_order(teachers: list[dict]):
        tm = _load_teachers_manage_module()
        mgmt_rank = {n: i + 1 for i, n in enumerate(getattr(tm, 'MGMT_ORDER', []) or [])}
        for t in teachers:
            name = getattr(tm, 'normalize_name')(t.get('name') or '')
            rank = mgmt_rank.get(name)
            if not rank:
                continue
            for r in (t.get('roles') or []):
                if getattr(tm, 'clean_dept')(r.get('department') or '') == '管理部':
                    r['order'] = rank

    def _role_key(t: dict, r: dict) -> str:
        return f"{t.get('id','')}::{str(r.get('department') or '')}::{str(r.get('position') or '')}"

    def _normalize_all_teachers(teachers: list[dict]) -> list[dict]:
        tm = _load_teachers_manage_module()
        teachers = getattr(tm, 'normalize_data')(teachers)
        teachers = getattr(tm, 'merge_teachers_by_name')(teachers)
        for t in teachers:
            getattr(tm, 'normalize_teacher_roles')(t)
            t['roles'] = sorted(t.get('roles') or [], key=lambda x: _safe_int(x.get('order')))
        _enforce_mgmt_order(teachers)
        for t in teachers:
            t['roles'] = sorted(t.get('roles') or [], key=lambda x: _safe_int(x.get('order')))

        def teacher_key(tt):
            orders = [_safe_int(r.get('order')) for r in (tt.get('roles') or [])]
            return (min(orders) if orders else 10**9, str(tt.get('name') or ''))

        teachers.sort(key=teacher_key)
        return teachers

    def _normalize_students(students: list[dict]) -> tuple[list[dict], list[str]]:
        sm = _load_students_manage_module()
        norm_line = getattr(sm, 'norm_line')

        if not isinstance(students, list):
            return [], ['students.json 顶层必须是数组(list)']

        # 规范字段
        for s in students:
            if not isinstance(s, dict):
                continue
            s['id'] = norm_line(str(s.get('id') or ''))
            s['name'] = norm_line(str(s.get('name') or '')).replace(' ', '')
            s['school'] = norm_line(str(s.get('school') or ''))
            s['major'] = norm_line(str(s.get('major') or ''))
            s['photo'] = norm_line(str(s.get('photo') or ''))

            year = s.get('year')
            if year is None or year == '':
                s['year'] = None
            else:
                try:
                    s['year'] = int(year)
                except Exception:
                    # 保留原值让 validate 报错更明确
                    pass

            admissions = s.get('admissions')
            if admissions is None:
                admissions = []
            if not isinstance(admissions, list):
                admissions = []

            cleaned = []
            for a in admissions:
                if not isinstance(a, dict):
                    continue
                img = norm_line(str(a.get('image') or ''))
                if not img:
                    continue
                cleaned.append({
                    'image': img,
                    'watermarked': bool(a.get('watermarked')),
                    'note': norm_line(str(a.get('note') or '')),
                })
            s['admissions'] = cleaned

        # 默认排序：year desc, name asc
        def key(ss):
            yy = ss.get('year')
            try:
                yy = int(yy) if yy is not None else 0
            except Exception:
                yy = 0
            return (-yy, str(ss.get('name') or ''))

        students = sorted(students, key=key)

        ok, errs = getattr(sm, 'validate_data')(students)
        return students, ([] if ok else errs)

    @app.get('/')
    def root():
        if is_authed():
            return redirect(url_for('admin_home'))
        return redirect(url_for('login'))

    @app.get('/login')
    def login():
        return render_template('login.html')

    @app.post('/login')
    def login_post():
        pwd = request.form.get('password') or ''
        expected = os.environ.get('ADMIN_PASSWORD') or ''
        if not expected:
            flash('服务器未设置 ADMIN_PASSWORD，无法登录。请在启动前设置环境变量。', 'error')
            return redirect(url_for('login'))

        if pwd != expected:
            flash('密码错误', 'error')
            return redirect(url_for('login'))

        session['admin_authed'] = True
        flash('已登录', 'ok')
        nxt = request.form.get('next') or request.args.get('next') or url_for('admin_home')
        return redirect(nxt)

    @app.after_request
    def _no_cache_for_admin(resp):
        try:
            p = request.path or ''
            if p.startswith('/admin') or p.startswith('/login'):
                resp.headers['Cache-Control'] = 'no-store'
        except Exception:
            pass
        return resp

    @app.get('/logout')
    def logout():
        session.clear()
        return redirect(url_for('login'))

    @app.get('/admin')
    @login_required
    def admin_home():
        teachers = load_json(DATA_TEACHERS)
        students = load_json(DATA_STUDENTS)
        portal_posts = _load_portal_posts()
        return render_template(
            'home.html',
            teacher_count=len(teachers),
            student_count=len(students),
            portal_count=len(portal_posts),
        )

    @app.get('/admin/portal')
    @login_required
    def portal_list_page():
        q = str(request.args.get('q') or '').strip().lower()
        posts = _load_portal_posts()

        def hit(p: dict) -> bool:
            if not q:
                return True
            hay = ' '.join([
                str(p.get('title') or ''),
                str(p.get('summary') or ''),
                str(p.get('category') or ''),
                ' '.join([str(t) for t in (p.get('tags') or [])]),
            ]).lower()
            return q in hay

        filtered = [p for p in posts if hit(p)]

        def sort_key(p: dict):
            pinned = 0 if p.get('pinned') else 1
            dt = p.get('publishedAt') or p.get('updatedAt') or ''
            return (pinned, str(dt))

        filtered.sort(key=sort_key, reverse=False)
        # pinned first, then time desc
        filtered = sorted(
            filtered,
            key=lambda p: (
                1 if not p.get('pinned') else 0,
                str(p.get('publishedAt') or p.get('updatedAt') or ''),
            ),
            reverse=False,
        )
        # reverse time within groups
        filtered = list(reversed(filtered))

        return render_template('portal_list.html', posts=filtered, q=q)

    @app.get('/admin/portal/new')
    @login_required
    def portal_new_page():
        post = {
            'id': '',
            'title': '',
            'category': '通知',
            'tags': [],
            'status': 'draft',
            'pinned': False,
            'coverImage': '',
            'shareUrl': '',
            'summary': '',
            'publishedAt': '',
            'updatedAt': '',
            'bodyHtml': '',
        }
        return render_template('portal_edit.html', post=post, is_new=True)

    @app.post('/admin/portal/new')
    @login_required
    def portal_create_post():
        posts = _load_portal_posts()
        pid = _make_id('p')
        now = _now_iso()

        status = str(request.form.get('status') or 'draft').strip().lower()
        status = status if status in ('draft', 'published') else 'draft'
        published_at = str(request.form.get('publishedAt') or '').strip()
        if status == 'published' and not published_at:
            published_at = now

        post = {
            'id': pid,
            'title': str(request.form.get('title') or '').strip(),
            'category': str(request.form.get('category') or '').strip(),
            'tags': _norm_tags(str(request.form.get('tags') or '')),
            'status': status,
            'pinned': bool(request.form.get('pinned')),
            'coverImage': str(request.form.get('coverImage') or '').strip(),
            'shareUrl': str(request.form.get('shareUrl') or '').strip(),
            'summary': str(request.form.get('summary') or '').strip(),
            'publishedAt': published_at,
            'updatedAt': now,
            'bodyHtml': str(request.form.get('bodyHtml') or '').strip(),
        }

        if not post['title']:
            flash('标题不能为空', 'error')
            return redirect(url_for('portal_new_page'))

        posts.append(post)
        _save_portal_posts(posts)
        flash('已创建', 'ok')
        return redirect(url_for('portal_edit_page', pid=pid))

    @app.get('/admin/portal/<pid>')
    @login_required
    def portal_edit_page(pid: str):
        posts = _load_portal_posts()
        post = _find_post(posts, pid)
        if not post:
            flash('文章不存在', 'error')
            return redirect(url_for('portal_list_page'))
        return render_template('portal_edit.html', post=post, is_new=False)

    @app.post('/admin/portal/<pid>')
    @login_required
    def portal_update_post(pid: str):
        posts = _load_portal_posts()
        post = _find_post(posts, pid)
        if not post:
            flash('文章不存在', 'error')
            return redirect(url_for('portal_list_page'))

        title = str(request.form.get('title') or '').strip()
        if not title:
            flash('标题不能为空', 'error')
            return redirect(url_for('portal_edit_page', pid=pid))

        status = str(request.form.get('status') or 'draft').strip().lower()
        status = status if status in ('draft', 'published') else 'draft'
        published_at = str(request.form.get('publishedAt') or '').strip()
        if status == 'published' and not published_at and not str(post.get('publishedAt') or '').strip():
            published_at = _now_iso()
        if not published_at:
            published_at = str(post.get('publishedAt') or '').strip()

        post['title'] = title
        post['category'] = str(request.form.get('category') or '').strip()
        post['tags'] = _norm_tags(str(request.form.get('tags') or ''))
        post['status'] = status
        post['pinned'] = bool(request.form.get('pinned'))
        post['coverImage'] = str(request.form.get('coverImage') or '').strip()
        post['shareUrl'] = str(request.form.get('shareUrl') or '').strip()
        post['summary'] = str(request.form.get('summary') or '').strip()
        post['publishedAt'] = published_at
        post['updatedAt'] = _now_iso()
        post['bodyHtml'] = str(request.form.get('bodyHtml') or '').strip()

        _save_portal_posts(posts)
        flash('已保存', 'ok')
        return redirect(url_for('portal_edit_page', pid=pid))

    @app.post('/admin/portal/<pid>/delete')
    @login_required
    def portal_delete_post(pid: str):
        posts = _load_portal_posts()
        before = len(posts)
        posts = [p for p in posts if str(p.get('id') or '') != str(pid)]
        if len(posts) == before:
            flash('文章不存在', 'error')
            return redirect(url_for('portal_list_page'))
        _save_portal_posts(posts)
        flash('已删除', 'ok')
        return redirect(url_for('portal_list_page'))

    @app.get('/admin/portal/<pid>/preview')
    @login_required
    def portal_preview(pid: str):
        posts = _load_portal_posts()
        post = _find_post(posts, pid)
        if not post:
            flash('文章不存在', 'error')
            return redirect(url_for('portal_list_page'))
        return render_template('portal_preview.html', post=post)

    @app.get('/admin/portal/<pid>/poster.png')
    @login_required
    def portal_poster(pid: str):
        """Generate a vertical poster PNG for WeChat public account."""
        if Image is None or ImageDraw is None or ImageFont is None:
            return jsonify({'ok': False, 'error': '缺少 Pillow，无法生成海报（请先安装 pillow）'}), 500

        if qrcode is None:
            # still allow poster without QR
            pass

        posts = _load_portal_posts()
        post = _find_post(posts, pid)
        if not post:
            return jsonify({'ok': False, 'error': 'post not found'}), 404

        theme = str(request.args.get('theme') or 'brand').strip().lower()
        theme = theme if theme in ('brand', 'minimal') else 'brand'

        # fixed portrait size, can be overridden
        try:
            w = int(request.args.get('w') or 1080)
            h = int(request.args.get('h') or 1920)
        except Exception:
            w, h = 1080, 1920
        w = max(720, min(1440, w))
        h = max(1280, min(2560, h))

        if theme == 'minimal':
            bg = Image.new('RGB', (w, h), (248, 248, 248))
            accent = (24, 24, 26)
            text_main = (18, 18, 20)
            text_muted = (110, 110, 118)
            brand = (24, 24, 26)
        else:
            bg = Image.new('RGB', (w, h), (250, 250, 252))
            accent = (102, 8, 116)
            text_main = (20, 20, 22)
            text_muted = (110, 110, 118)
            brand = (102, 8, 116)
        draw = ImageDraw.Draw(bg)

        pad = int(w * 0.065)
        x0 = pad
        x1 = w - pad
        y = pad

        font_title = _pick_cjk_font(int(w * 0.056))
        font_meta = _pick_cjk_font(int(w * 0.028))
        font_body = _pick_cjk_font(int(w * 0.032))
        font_small = _pick_cjk_font(int(w * 0.024))

        def text_bbox(txt: str, font):
            if not txt:
                return (0, 0, 0, 0)
            return draw.textbbox((0, 0), txt, font=font)

        # top accent + card
        card_r = int(w * 0.04)
        card = (x0, y, x1, h - pad)
        draw.rounded_rectangle(card, radius=card_r, fill=(255, 255, 255))

        # accent bar
        accent_h = max(10, int(w * 0.012))
        ax0 = x0 + int(w * 0.05)
        ax1 = x1 - int(w * 0.05)
        ay0 = y + int(w * 0.05)
        draw.rounded_rectangle((ax0, ay0, ax1, ay0 + accent_h), radius=accent_h // 2, fill=accent)
        y = ay0 + accent_h + int(w * 0.04)

        # shift content inward inside card
        x0c = x0 + int(w * 0.05)
        x1c = x1 - int(w * 0.05)

        title = str(post.get('title') or '未命名').strip()
        cat = str(post.get('category') or '').strip()
        date_s = str(post.get('publishedAt') or post.get('updatedAt') or '').strip()
        if date_s and 'T' in date_s:
            date_s = date_s.split('T', 1)[0]

        # title wrap (max 2 lines)
        title_lines = _wrap_text(draw, title, font_title, x1c - x0c)
        title_lines = title_lines[:2] if title_lines else ['未命名']
        for ln in title_lines:
            draw.text((x0c, y), ln, font=font_title, fill=text_main)
            y += (text_bbox(ln, font_title)[3] - text_bbox(ln, font_title)[1]) + int(w * 0.012)

        meta_parts = []
        if cat:
            meta_parts.append(cat)
        if date_s:
            meta_parts.append(date_s)
        meta = ' · '.join(meta_parts) if meta_parts else '润德信息门户'
        draw.text((x0c, y), meta, font=font_meta, fill=text_muted)
        y += (text_bbox(meta, font_meta)[3] - text_bbox(meta, font_meta)[1]) + int(w * 0.03)

        # cover image (optional)
        cover = str(post.get('coverImage') or '').strip()
        if cover:
            im = _load_image_any(cover)
            if im is not None:
                cover_h = int(h * 0.28)
                cover_w = x1c - x0c
                # crop to fill
                src_w, src_h = im.size
                target_ratio = cover_w / cover_h
                src_ratio = src_w / src_h if src_h else 1
                if src_ratio > target_ratio:
                    # wider
                    new_h = src_h
                    new_w = int(new_h * target_ratio)
                    left = (src_w - new_w) // 2
                    im = im.crop((left, 0, left + new_w, new_h))
                else:
                    new_w = src_w
                    new_h = int(new_w / target_ratio)
                    top = (src_h - new_h) // 2
                    im = im.crop((0, top, new_w, top + new_h))
                im = im.resize((cover_w, cover_h))

                # rounded mask
                r = int(w * 0.03)
                mask = Image.new('L', (cover_w, cover_h), 0)
                md = ImageDraw.Draw(mask)
                md.rounded_rectangle((0, 0, cover_w, cover_h), radius=r, fill=255)
                bg.paste(im, (x0c, y), mask)
                y += cover_h + int(w * 0.03)

        # body text
        body_html = str(post.get('bodyHtml') or '').strip()
        body_text = _html_to_text_simple(body_html)
        if not body_text:
            body_text = str(post.get('summary') or '').strip()

        try:
            max_lines_arg = request.args.get('max_lines')
            max_lines_override = int(max_lines_arg) if max_lines_arg else None
        except Exception:
            max_lines_override = None

        max_body_h = (h - pad) - y - int(w * 0.22)
        line_h = int(w * 0.048)
        max_lines = max(6, max_body_h // line_h)
        if max_lines_override is not None:
            max_lines = max(4, min(40, max_lines_override))

        lines = _wrap_text(draw, body_text, font_body, x1c - x0c)
        # trim leading empty lines
        while lines and not lines[0].strip():
            lines.pop(0)

        truncated = False
        if len(lines) > max_lines:
            lines = lines[:max_lines]
            truncated = True

        for ln in lines:
            if not ln:
                y += int(line_h * 0.65)
                continue
            draw.text((x0c, y), ln, font=font_body, fill=(35, 35, 38) if theme == 'brand' else text_main)
            y += line_h

        if truncated:
            ell = '（更多内容请在官网信息门户查看）'
            draw.text((x0c, y + int(w * 0.01)), ell, font=font_small, fill=brand)
            y += int(w * 0.06)

        # footer: optional QR + caption
        share_url = str(request.args.get('url') or post.get('shareUrl') or '').strip()
        show_qr = str(request.args.get('qr') or '1').strip().lower() not in ('0', 'false', 'no')
        qr_size = int(w * 0.22)
        footer_y = (h - pad) - int(w * 0.16)
        if show_qr and share_url:
            qr_img = _make_qr(share_url, size=qr_size)
            if qr_img is not None:
                qx = x1c - qr_size
                qy = footer_y
                # white background for QR
                r = int(w * 0.018)
                draw.rounded_rectangle((qx - int(w * 0.012), qy - int(w * 0.012), qx + qr_size + int(w * 0.012), qy + qr_size + int(w * 0.012)), radius=r, fill=(255, 255, 255), outline=(0, 0, 0, 18))
                bg.paste(qr_img, (qx, qy))

                cap = '扫码查看原文'
                cap_w = draw.textlength(cap, font=font_small) if hasattr(draw, 'textlength') else text_bbox(cap, font_small)[2]
                draw.text((qx + (qr_size - cap_w) / 2, qy + qr_size + int(w * 0.01)), cap, font=font_small, fill=text_muted)

        footer = '兰州润德艺术学校 · 信息门户'
        draw.text((x0c, h - pad - int(w * 0.08)), footer, font=font_small, fill=text_muted)
        if share_url:
            # show short hostname
            try:
                host = urllib.parse.urlparse(share_url).netloc
            except Exception:
                host = ''
            if host:
                draw.text((x0c, h - pad - int(w * 0.045)), host, font=font_small, fill=brand)

        out = io.BytesIO()
        bg.save(out, format='PNG', optimize=True)
        out.seek(0)

        safe_title = _safe_filename(title)
        fname = f'poster_{pid}_{safe_title}.png'
        return send_file(out, as_attachment=True, download_name=fname, mimetype='image/png')

    @app.post('/admin/portal/upload-image')
    @login_required
    def portal_upload_image():
        f = request.files.get('image')
        if not f or not getattr(f, 'filename', None):
            return jsonify({'ok': False, 'error': 'missing image'}), 400

        orig = str(f.filename or '')
        ext = (Path(orig).suffix or '').lower()
        if ext not in {'.jpg', '.jpeg', '.png', '.webp'}:
            return jsonify({'ok': False, 'error': 'unsupported image type'}), 400

        yyyymm = datetime.datetime.now().strftime('%Y%m')
        dest_dir = ROOT / 'assets' / 'portal' / yyyymm
        dest_dir.mkdir(parents=True, exist_ok=True)

        safe_base = _safe_filename(Path(orig).stem)
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        name = f'{ts}_{safe_base}{ext}'
        dest = dest_dir / name

        try:
            f.stream.seek(0)
        except Exception:
            pass
        with dest.open('wb') as w:
            shutil.copyfileobj(f.stream, w)

        rel = f'assets/portal/{yyyymm}/{name}'
        return jsonify({'ok': True, 'path': rel})

    @app.get('/site/<path:relpath>')
    @login_required
    def site_file(relpath: str):
        """Expose selected project files for admin preview only."""
        try:
            rp = _norm_json_path(relpath)
        except Exception:
            return jsonify({'ok': False, 'error': 'invalid path'}), 400

        if not (rp.startswith('assets/portal/') or rp == 'assets/portal'):
            return jsonify({'ok': False, 'error': 'forbidden'}), 403

        p = ROOT / rp
        if not p.exists() or not p.is_file():
            return jsonify({'ok': False, 'error': 'not found'}), 404

        return send_file(p)

    @app.get('/admin/import')
    @login_required
    def bulk_import_page():
        return render_template('bulk_import.html', result=None)

    @app.get('/admin/import/templates/<kind>.xlsx')
    @login_required
    def bulk_import_excel_template(kind: str):
        if load_workbook is None:
            flash('缺少 openpyxl，无法生成 Excel 模板（请先安装 openpyxl）', 'error')
            return redirect(url_for('bulk_import_page'))

        # openpyxl Workbook is available via import above
        from openpyxl import Workbook  # type: ignore

        kind = (kind or '').strip().lower()
        if kind not in ('teachers', 'students'):
            return jsonify({'ok': False, 'error': 'unknown template kind'}), 404

        wb = Workbook()
        ws = wb.active

        if kind == 'teachers':
            ws.title = 'teachers'
            ws.append(['id', 'name', 'photo', 'shortSummary', 'bio', 'achievements', 'department', 'position', 'order'])
            ws.append(['t_001', '张老师', 'teachers/photos/zhang.jpg', '一句话简介', '详细介绍', '奖项A|奖项B', '声乐组', '声乐教师', 10])
            ws.append(['t_001', '张老师', '', '', '', '', '管理部', '（待设置）', 1])
        else:
            ws.title = 'students'
            ws.append(['id', 'name', 'school', 'major', 'year', 'photo', 'admission_image', 'admission_note', 'admission_watermarked'])
            ws.append(['s_001', '李同学', '上海音乐学院', '钢琴演奏', 2026, 'students/photos/li.png', 'students/admissions/li_offer.png', '录取截图', '是'])
            ws.append(['s_001', '李同学', '上海音乐学院', '钢琴演奏', 2026, 'students/photos/li.png', '', '', ''])

        # Add a README sheet with strict rules
        readme = wb.create_sheet('README')
        if kind == 'teachers':
            lines = [
                '教师 Excel 模板说明',
                '',
                '必填列：id, name',
                '可选列：photo, shortSummary, bio, achievements, department, position, order',
                '规则：',
                '- 每行可代表一个岗位；同一 id 多行会合并为同一老师并累计 roles。',
                '- department/position 必须同时填写才会生成岗位。',
                '- photo 推荐写 teachers/photos/xxx.jpg（若同时上传文件夹，可自动复制到项目 photos/）。',
                '- achievements 用 | 分隔多个条目。',
            ]
        else:
            lines = [
                '学生 Excel 模板说明',
                '',
                '必填列：id, name, school, major, year',
                '可选列：photo, admission_image, admission_note, admission_watermarked',
                '规则：',
                '- 同一 id 多行会合并为同一学生，并累计 admissions。',
                '- photo 必须写 students/photos/xxx.jpg',
                '- admission_image 必须写 students/admissions/xxx.jpg',
                '- admission_watermarked 支持：是/否、1/0、true/false',
            ]
        for i, line in enumerate(lines, start=1):
            readme.cell(row=i, column=1, value=line)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f'{kind}_import_template.xlsx'
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @app.post('/admin/import/teachers')
    @login_required
    def bulk_import_teachers():
        mode = str(request.form.get('mode') or 'merge').strip()
        mode = mode if mode in ('merge', 'replace') else 'merge'
        dry_run = bool(request.form.get('dry_run'))

        files = request.files.getlist('files')
        pack = _normalize_webkitdirectory_upload(files)
        m = pack['files']
        errors: list[str] = []
        warnings: list[str] = []
        created = updated = files_copied = 0
        records = 0

        try:
            incoming = _read_json_from_upload(m, 'teachers/teachers.json')
        except FileNotFoundError:
            errors.append('缺少 teachers/teachers.json')
            incoming = []
        except Exception as e:
            errors.append(f'teachers/teachers.json 解析失败：{e}')
            incoming = []

        if incoming and not isinstance(incoming, list):
            errors.append('teachers/teachers.json 顶层必须是数组(list)')
            incoming = []

        normalized: list[dict] = []
        if incoming:
            for i, t in enumerate(incoming):
                if not isinstance(t, dict):
                    errors.append(f'teachers.json 项 {i} 不是对象')
                    continue
                tid = str(t.get('id') or '').strip()
                name = str(t.get('name') or '').strip()
                if not tid or not name:
                    errors.append(f'teachers.json 项 {i} 缺少 id/name')
                    continue

                tt = dict(t)
                # photo mapping: teachers/photos/x -> photos/x
                if 'photo' in tt and tt.get('photo'):
                    try:
                        p = _norm_json_path(tt.get('photo'))
                    except Exception as e:
                        errors.append(f'teachers.json {tid} photo 非法：{e}')
                        p = ''
                    if p:
                        if not (p.startswith('teachers/photos/') or p.startswith('photos/')):
                            errors.append(f'teachers.json {tid} photo 必须写 teachers/photos/...')
                        else:
                            fn = Path(p).name
                            tt['photo'] = f'photos/{fn}'
                            # locate source file in upload
                            src1 = f'teachers/photos/{fn}'
                            src2 = f'photos/{fn}'
                            src = src1 if src1 in m else (src2 if src2 in m else '')
                            if not src:
                                errors.append(f'teachers.json {tid} 缺少照片文件：teachers/photos/{fn}')
                            elif not _is_allowed_image(fn):
                                errors.append(f'teachers.json {tid} 照片后缀不支持：{fn}')
                            else:
                                if not dry_run and not errors:
                                    _save_uploaded_file(m, src, ROOT / 'photos' / fn)
                                files_copied += 1

                normalized.append(tt)

        records = len(normalized)

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'teachers',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': records,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        # merge/replace into data
        existing = load_json(DATA_TEACHERS)
        existing, created, updated = _merge_by_id(existing, normalized, mode=mode)
        existing = _normalize_all_teachers(existing)
        _enforce_mgmt_order(existing)

        if not dry_run:
            write_json(DATA_TEACHERS, existing)

        return render_template('bulk_import.html', result={
            'ok': True,
            'kind': 'teachers',
            'mode': mode + (' (dry-run)' if dry_run else ''),
            'records': records,
            'created': created,
            'updated': updated,
            'files_copied': files_copied,
            'errors': [],
            'warnings': warnings,
        })

    @app.post('/admin/import/students')
    @login_required
    def bulk_import_students():
        mode = str(request.form.get('mode') or 'merge').strip()
        mode = mode if mode in ('merge', 'replace') else 'merge'
        dry_run = bool(request.form.get('dry_run'))

        files = request.files.getlist('files')
        pack = _normalize_webkitdirectory_upload(files)
        m = pack['files']
        errors: list[str] = []
        warnings: list[str] = []
        created = updated = files_copied = 0
        records = 0

        try:
            incoming = _read_json_from_upload(m, 'students/students.json')
        except FileNotFoundError:
            errors.append('缺少 students/students.json')
            incoming = []
        except Exception as e:
            errors.append(f'students/students.json 解析失败：{e}')
            incoming = []

        if incoming and not isinstance(incoming, list):
            errors.append('students/students.json 顶层必须是数组(list)')
            incoming = []

        normalized: list[dict] = []
        if incoming:
            for i, s in enumerate(incoming):
                if not isinstance(s, dict):
                    errors.append(f'students.json 项 {i} 不是对象')
                    continue
                sid = str(s.get('id') or '').strip()
                name = str(s.get('name') or '').strip()
                if not sid or not name:
                    errors.append(f'students.json 项 {i} 缺少 id/name')
                    continue

                ss = dict(s)

                # photo must be students/photos/x
                if 'photo' in ss and ss.get('photo'):
                    try:
                        p = _norm_json_path(ss.get('photo'))
                    except Exception as e:
                        errors.append(f'students.json {sid} photo 非法：{e}')
                        p = ''
                    if p:
                        if not p.startswith('students/photos/'):
                            errors.append(f'students.json {sid} photo 必须写 students/photos/...')
                        else:
                            fn = Path(p).name
                            if not _is_allowed_image(fn):
                                errors.append(f'students.json {sid} photo 后缀不支持：{fn}')
                            elif p not in m:
                                errors.append(f'students.json {sid} 缺少照片文件：{p}')
                            else:
                                if not dry_run and not errors:
                                    _save_uploaded_file(m, p, ROOT / 'students' / 'photos' / fn)
                                files_copied += 1
                            ss['photo'] = f'students/photos/{fn}'

                admissions = ss.get('admissions')
                if admissions is None:
                    admissions = []
                if not isinstance(admissions, list):
                    errors.append(f'students.json {sid} admissions 必须是数组(list)')
                    admissions = []

                new_adm = []
                for ai, a in enumerate(admissions):
                    if not isinstance(a, dict):
                        errors.append(f'students.json {sid} admissions[{ai}] 不是对象')
                        continue
                    img = a.get('image')
                    if not img:
                        errors.append(f'students.json {sid} admissions[{ai}] 缺少 image')
                        continue
                    try:
                        p = _norm_json_path(img)
                    except Exception as e:
                        errors.append(f'students.json {sid} admissions[{ai}].image 非法：{e}')
                        continue
                    if not p.startswith('students/admissions/'):
                        errors.append(f'students.json {sid} admissions[{ai}].image 必须写 students/admissions/...')
                        continue
                    fn = Path(p).name
                    if not _is_allowed_image(fn):
                        errors.append(f'students.json {sid} admissions[{ai}] 后缀不支持：{fn}')
                        continue
                    if p not in m:
                        errors.append(f'students.json {sid} 缺少 admissions 文件：{p}')
                        continue
                    if not dry_run and not errors:
                        _save_uploaded_file(m, p, ROOT / 'students' / 'admissions' / fn)
                    files_copied += 1

                    new_adm.append({
                        'image': f'students/admissions/{fn}',
                        'watermarked': bool(a.get('watermarked')),
                        'note': a.get('note') or '',
                    })
                ss['admissions'] = new_adm
                normalized.append(ss)

        records = len(normalized)

        # validate using students/manage.py rules
        if not errors:
            sm = _load_students_manage_module()
            ok, errs = getattr(sm, 'validate_data')(normalized)
            if not ok:
                errors.extend(errs)

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'students',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': records,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        existing = load_json(DATA_STUDENTS)
        existing, created, updated = _merge_by_id(existing, normalized, mode=mode)
        # keep a stable order: year desc, name asc
        existing.sort(key=lambda x: (-_safe_int(x.get('year'), default=0), str(x.get('name') or '')))
        if not dry_run:
            write_json(DATA_STUDENTS, existing)

        return render_template('bulk_import.html', result={
            'ok': True,
            'kind': 'students',
            'mode': mode + (' (dry-run)' if dry_run else ''),
            'records': records,
            'created': created,
            'updated': updated,
            'files_copied': files_copied,
            'errors': [],
            'warnings': warnings,
        })

    @app.post('/admin/import/teachers-excel')
    @login_required
    def bulk_import_teachers_excel():
        mode = str(request.form.get('mode') or 'merge').strip()
        mode = mode if mode in ('merge', 'replace') else 'merge'
        dry_run = bool(request.form.get('dry_run'))

        f = request.files.get('excel')
        errors: list[str] = []
        warnings: list[str] = []
        created = updated = 0
        records = 0
        files_copied = 0

        # optional folder upload for photos
        folder_files = request.files.getlist('files')
        folder_pack = _normalize_webkitdirectory_upload(folder_files)
        folder_map = folder_pack['files']

        if not f or not getattr(f, 'filename', ''):
            errors.append('请上传 .xlsx 文件')
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'teachers-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': 0,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        try:
            rows = _read_xlsx_rows(f, required_cols=['id', 'name'])
        except Exception as e:
            errors.append(f'Excel 解析失败：{e}')
            rows = []

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'teachers-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': 0,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        # aggregate: one row can represent one role; same id merges roles
        agg: dict[str, dict] = {}
        for i, r in enumerate(rows):
            tid = str(r.get('id') or '').strip()
            name = str(r.get('name') or '').strip()
            if not tid or not name:
                errors.append(f'第 {i+2} 行缺少 id/name')
                continue

            t = agg.get(tid) or {
                'id': tid,
                'name': name,
                'photo': '',
                'shortSummary': '',
                'bio': '',
                'achievements': [],
                'roles': [],
            }
            t['name'] = name

            for k in ('photo', 'shortSummary', 'bio'):
                if k in r and r.get(k) is not None:
                    t[k] = str(r.get(k) or '').strip()

            # achievements: split by | or newline
            if 'achievements' in r and r.get('achievements') is not None:
                raw = str(r.get('achievements') or '').strip()
                if raw:
                    parts = [x.strip() for x in re.split(r'[\n\|]+', raw) if x and x.strip()]
                    t['achievements'] = parts

            dept = str(r.get('department') or '').strip() if 'department' in r else ''
            pos = str(r.get('position') or '').strip() if 'position' in r else ''
            if dept and pos:
                try:
                    order = _safe_int(r.get('order'), default=10**9)
                except Exception:
                    order = 10**9
                t.setdefault('roles', []).append({'department': dept, 'position': pos, 'order': order})
            elif dept or pos:
                warnings.append(f'第 {i+2} 行岗位信息不完整（department/position 需同时填写），已忽略该岗位')

            agg[tid] = t

        # normalize teacher photo: accept photos/xxx or teachers/photos/xxx
        normalized: list[dict] = []
        for tid, t in agg.items():
            photo = str(t.get('photo') or '').strip()
            if photo:
                try:
                    p = _norm_json_path(photo)
                except Exception as e:
                    errors.append(f'{tid} photo 非法：{e}')
                    p = ''
                if p:
                    if p.startswith('teachers/photos/') or p.startswith('photos/'):
                        fn = Path(p).name
                        t['photo'] = f'photos/{fn}'

                        # If folder uploaded, try to copy. Otherwise, only warn if missing.
                        if folder_map:
                            src = ''
                            # prefer exact paths
                            cand1 = f'teachers/photos/{fn}'
                            cand2 = f'photos/{fn}'
                            if cand1 in folder_map:
                                src = cand1
                            elif cand2 in folder_map:
                                src = cand2
                            else:
                                src = _find_upload_by_basename(folder_map, fn) or ''
                                # ambiguous matches
                                if not src and any(Path(k).name == fn for k in folder_map.keys()):
                                    errors.append(f'{tid} photo 文件名重复，无法确定用哪个：{fn}（请保证唯一）')

                            if not src:
                                errors.append(f'{tid} photo 在上传文件夹中找不到：{p}')
                            else:
                                if not dry_run:
                                    _save_uploaded_file(folder_map, src, ROOT / 'photos' / fn)
                                    files_copied += 1
                        else:
                            if not (ROOT / 'photos' / fn).exists():
                                warnings.append(f'{tid} photo 文件不存在于项目：photos/{fn}（可在 Excel 表单同时上传文件夹自动复制）')
                    else:
                        errors.append(f'{tid} photo 必须写 photos/... 或 teachers/photos/...')

            normalized.append(t)

        records = len(normalized)

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'teachers-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': records,
                'created': 0,
                'updated': 0,
                'files_copied': files_copied,
                'errors': errors,
                'warnings': warnings,
            })

        existing = load_json(DATA_TEACHERS)
        existing, created, updated = _merge_by_id(existing, normalized, mode=mode)
        existing = _normalize_all_teachers(existing)
        _enforce_mgmt_order(existing)
        if not dry_run:
            write_json(DATA_TEACHERS, existing)

        return render_template('bulk_import.html', result={
            'ok': True,
            'kind': 'teachers-excel',
            'mode': mode + (' (dry-run)' if dry_run else ''),
            'records': records,
            'created': created,
            'updated': updated,
            'files_copied': files_copied,
            'errors': [],
            'warnings': warnings,
        })

    @app.post('/admin/import/students-excel')
    @login_required
    def bulk_import_students_excel():
        mode = str(request.form.get('mode') or 'merge').strip()
        mode = mode if mode in ('merge', 'replace') else 'merge'
        dry_run = bool(request.form.get('dry_run'))

        f = request.files.get('excel')
        errors: list[str] = []
        warnings: list[str] = []
        created = updated = 0
        records = 0
        files_copied = 0

        # optional folder upload for photos/admissions
        folder_files = request.files.getlist('files')
        folder_pack = _normalize_webkitdirectory_upload(folder_files)
        folder_map = folder_pack['files']

        if not f or not getattr(f, 'filename', ''):
            errors.append('请上传 .xlsx 文件')
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'students-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': 0,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        try:
            rows = _read_xlsx_rows(f, required_cols=['id', 'name', 'school', 'major', 'year'])
        except Exception as e:
            errors.append(f'Excel 解析失败：{e}')
            rows = []

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'students-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': 0,
                'created': 0,
                'updated': 0,
                'files_copied': 0,
                'errors': errors,
                'warnings': warnings,
            })

        agg: dict[str, dict] = {}
        for i, r in enumerate(rows):
            sid = str(r.get('id') or '').strip()
            name = str(r.get('name') or '').strip()
            school = str(r.get('school') or '').strip()
            major = str(r.get('major') or '').strip()
            year_raw = r.get('year')
            try:
                year = int(year_raw) if year_raw is not None and str(year_raw).strip() != '' else None
            except Exception:
                year = None

            if not sid or not name or not school or not major or year is None:
                errors.append(f'第 {i+2} 行缺少必要字段（id/name/school/major/year）')
                continue

            s = agg.get(sid) or {
                'id': sid,
                'name': name,
                'school': school,
                'major': major,
                'year': year,
                'photo': '',
                'admissions': [],
            }
            s['name'] = name
            s['school'] = school
            s['major'] = major
            s['year'] = year

            if 'photo' in r and r.get('photo') is not None:
                s['photo'] = str(r.get('photo') or '').strip()

            adm_img = str(r.get('admission_image') or '').strip() if 'admission_image' in r else ''
            adm_note = str(r.get('admission_note') or '').strip() if 'admission_note' in r else ''
            adm_wm = _boolish(r.get('admission_watermarked')) if 'admission_watermarked' in r else False
            if adm_img:
                try:
                    p = _norm_json_path(adm_img)
                except Exception as e:
                    errors.append(f'{sid} admissions 第 {i+2} 行 image 非法：{e}')
                    p = ''
                if p:
                    if not p.startswith('students/admissions/'):
                        errors.append(f'{sid} admissions image 必须写 students/admissions/...')
                    else:
                        fn = Path(p).name
                        if not _is_allowed_image(fn):
                            errors.append(f'{sid} admissions 后缀不支持：{fn}')
                        else:
                            if folder_map:
                                src = ''
                                if p in folder_map:
                                    src = p
                                else:
                                    src = _find_upload_by_basename(folder_map, fn) or ''
                                    if not src and any(Path(k).name == fn for k in folder_map.keys()):
                                        errors.append(f'{sid} admissions 文件名重复，无法确定用哪个：{fn}（请保证唯一）')
                                if not src:
                                    errors.append(f'{sid} admissions 在上传文件夹中找不到：{p}')
                                else:
                                    if not dry_run:
                                        _save_uploaded_file(folder_map, src, ROOT / 'students' / 'admissions' / fn)
                                        files_copied += 1
                            else:
                                if not (ROOT / 'students' / 'admissions' / fn).exists():
                                    warnings.append(f'{sid} admissions 文件不存在于项目：students/admissions/{fn}（可在 Excel 表单同时上传文件夹自动复制）')
                            s.setdefault('admissions', []).append({'image': f'students/admissions/{fn}', 'watermarked': adm_wm, 'note': adm_note})
            agg[sid] = s

        normalized: list[dict] = []
        for sid, s in agg.items():
            photo = str(s.get('photo') or '').strip()
            if photo:
                try:
                    p = _norm_json_path(photo)
                except Exception as e:
                    errors.append(f'{sid} photo 非法：{e}')
                    p = ''
                if p:
                    if not p.startswith('students/photos/'):
                        errors.append(f'{sid} photo 必须写 students/photos/...')
                    else:
                        fn = Path(p).name
                        if not _is_allowed_image(fn):
                            errors.append(f'{sid} photo 后缀不支持：{fn}')
                        else:
                            if folder_map:
                                src = ''
                                if p in folder_map:
                                    src = p
                                else:
                                    src = _find_upload_by_basename(folder_map, fn) or ''
                                    if not src and any(Path(k).name == fn for k in folder_map.keys()):
                                        errors.append(f'{sid} photo 文件名重复，无法确定用哪个：{fn}（请保证唯一）')
                                if not src:
                                    errors.append(f'{sid} photo 在上传文件夹中找不到：{p}')
                                else:
                                    if not dry_run:
                                        _save_uploaded_file(folder_map, src, ROOT / 'students' / 'photos' / fn)
                                        files_copied += 1
                            else:
                                if not (ROOT / 'students' / 'photos' / fn).exists():
                                    warnings.append(f'{sid} photo 文件不存在于项目：students/photos/{fn}（可在 Excel 表单同时上传文件夹自动复制）')
                            s['photo'] = f'students/photos/{fn}'
            normalized.append(s)

        records = len(normalized)

        if not errors:
            sm = _load_students_manage_module()
            ok, errs = getattr(sm, 'validate_data')(normalized)
            if not ok:
                errors.extend(errs)

        if errors:
            return render_template('bulk_import.html', result={
                'ok': False,
                'kind': 'students-excel',
                'mode': mode + (' (dry-run)' if dry_run else ''),
                'records': records,
                'created': 0,
                'updated': 0,
                'files_copied': files_copied,
                'errors': errors,
                'warnings': warnings,
            })

        existing = load_json(DATA_STUDENTS)
        existing, created, updated = _merge_by_id(existing, normalized, mode=mode)
        existing.sort(key=lambda x: (-_safe_int(x.get('year'), default=0), str(x.get('name') or '')))
        if not dry_run:
            write_json(DATA_STUDENTS, existing)

        return render_template('bulk_import.html', result={
            'ok': True,
            'kind': 'students-excel',
            'mode': mode + (' (dry-run)' if dry_run else ''),
            'records': records,
            'created': created,
            'updated': updated,
            'files_copied': files_copied,
            'errors': [],
            'warnings': warnings,
        })

    @app.get('/admin/teachers')
    @login_required
    def teachers_list():
        teachers = load_json(DATA_TEACHERS)

        teacher_count = len(teachers)

        # 统一管理页：按部门聚合岗位卡片（同一老师跨部门会出现多张卡片）
        dept_map: dict[str, list[dict]] = {}
        for t in teachers:
            for r in (t.get('roles') or []):
                dept = str(r.get('department') or '').strip()
                pos = str(r.get('position') or '').strip()
                if not dept or not pos:
                    continue
                dept_map.setdefault(dept, []).append({
                    'roleKey': _role_key(t, r),
                    'teacherId': t.get('id'),
                    'name': t.get('name'),
                    'photo': t.get('photo') or '',
                    'shortSummary': t.get('shortSummary') or '',
                    'department': dept,
                    'position': pos,
                    'order': _safe_int(r.get('order')),
                })

        # 部门顺序：按该部门最小 order；舞蹈部置底
        def dept_key(item):
            dept, roles = item
            min_order = min((x.get('order', 10**9) for x in roles), default=10**9)
            dance_last = 1 if dept == '舞蹈部' else 0
            return (dance_last, min_order, dept)

        depts = []
        for dept, roles in dept_map.items():
            roles.sort(key=lambda x: (_safe_int(x.get('order')), str(x.get('name') or '')))
            depts.append({'department': dept, 'roles': roles})

        depts.sort(key=lambda d: dept_key((d['department'], d['roles'])))

        role_card_count = sum(len(d.get('roles') or []) for d in depts)
        return render_template('teachers_manage.html', depts=depts, teacher_count=teacher_count, role_card_count=role_card_count)

    @app.get('/admin/teachers/departments')
    @login_required
    def teachers_departments():
        # 兼容旧入口：统一跳转到 /admin/teachers（已合并）
        return redirect(url_for('teachers_list'))

    @app.post('/admin/teachers/departments/order')
    @login_required
    def teachers_departments_order():
        payload = request.get_json(silent=True) or {}
        dept = str(payload.get('department') or '').strip()
        role_keys = payload.get('roleKeys') or []
        if not dept or not isinstance(role_keys, list) or not role_keys:
            return jsonify({'ok': False, 'error': 'bad payload'}), 400

        teachers = load_json(DATA_TEACHERS)

        # 收集该部门当前 order 的最小值作为基准，保持部门整体排序区间不剧烈变化
        cur_orders = []
        for t in teachers:
            for r in (t.get('roles') or []):
                if str(r.get('department') or '').strip() == dept:
                    cur_orders.append(_safe_int(r.get('order')))
        base = min(cur_orders) if cur_orders else 1

        wanted = [str(k) for k in role_keys]
        wanted_set = set(wanted)
        updated = 0

        # 写回顺序（只影响该部门的 role）
        idx = 0
        for rk in wanted:
            for t in teachers:
                if not t.get('id'):
                    continue
                for r in (t.get('roles') or []):
                    if str(r.get('department') or '').strip() != dept:
                        continue
                    if _role_key(t, r) == rk:
                        r['order'] = int(base + idx)
                        idx += 1
                        updated += 1
                        break

        # 未出现在 roleKeys 里的同部门条目，按原顺序跟在后面（防止丢失）
        rest = []
        for t in teachers:
            for r in (t.get('roles') or []):
                if str(r.get('department') or '').strip() != dept:
                    continue
                if _role_key(t, r) not in wanted_set:
                    rest.append((_safe_int(r.get('order')), _role_key(t, r), r))
        rest.sort(key=lambda x: x[0])
        for _, _, r in rest:
            r['order'] = int(base + idx)
            idx += 1
            updated += 1

        for t in teachers:
            t['roles'] = sorted(t.get('roles') or [], key=lambda x: _safe_int(x.get('order')))

        write_json(DATA_TEACHERS, teachers)
        return jsonify({'ok': True, 'updated': updated, 'base': base})

    @app.post('/admin/teachers/departments/add-random')
    @login_required
    def teachers_departments_add_random():
        payload = request.get_json(silent=True) or {}
        dept = str(payload.get('department') or '').strip()
        if not dept:
            return jsonify({'ok': False, 'error': 'missing department'}), 400

        teachers = load_json(DATA_TEACHERS)

        # 只从“当前部门还没有岗位”的老师里随机挑一个，避免同部门重复出现同一个人
        candidates: list[dict] = []
        for t in teachers:
            rs = t.get('roles') or []
            has_dept = any(str(r.get('department') or '').strip() == dept for r in rs)
            if not has_dept:
                candidates.append(t)

        if not candidates:
            return jsonify({'ok': False, 'error': 'no available teacher for this department'}), 400

        # 该部门的新 order 放到末尾
        max_order = 0
        for t in teachers:
            for r in (t.get('roles') or []):
                if str(r.get('department') or '').strip() == dept:
                    max_order = max(max_order, _safe_int(r.get('order'), default=0))

        t = random.choice(candidates)
        new_role = {
            'department': dept,
            'position': '（待设置）',
            'order': max_order + 1,
        }
        t.setdefault('roles', []).append(new_role)

        write_json(DATA_TEACHERS, teachers)

        return jsonify({
            'ok': True,
            'role': {
                'roleKey': _role_key(t, new_role),
                'teacherId': t.get('id'),
                'name': t.get('name'),
                'department': dept,
                'position': new_role['position'],
                'order': new_role['order'],
            }
        })

    @app.get('/admin/teachers/departments/available')
    @login_required
    def teachers_departments_available():
        dept = str(request.args.get('department') or '').strip()
        q = str(request.args.get('q') or '').strip().lower()
        if not dept:
            return jsonify({'ok': False, 'error': 'missing department'}), 400

        teachers = load_json(DATA_TEACHERS)
        out: list[dict] = []

        for t in teachers:
            tid = str(t.get('id') or '').strip()
            if not tid:
                continue
            name = str(t.get('name') or '').strip()
            rs = t.get('roles') or []
            has_dept = any(str(r.get('department') or '').strip() == dept for r in rs)
            if has_dept:
                continue
            if q and q not in name.lower() and q not in tid.lower():
                continue
            out.append({'id': tid, 'name': name})

        out.sort(key=lambda x: (x.get('name') or '', x.get('id') or ''))
        return jsonify({'ok': True, 'teachers': out})

    @app.post('/admin/teachers/departments/add-existing')
    @login_required
    def teachers_departments_add_existing():
        payload = request.get_json(silent=True) or {}
        dept = str(payload.get('department') or '').strip()
        teacher_id = str(payload.get('teacherId') or '').strip()
        if not dept:
            return jsonify({'ok': False, 'error': 'missing department'}), 400
        if not teacher_id:
            return jsonify({'ok': False, 'error': 'missing teacherId'}), 400

        teachers = load_json(DATA_TEACHERS)
        t = next((x for x in teachers if str(x.get('id') or '').strip() == teacher_id), None)
        if not t:
            return jsonify({'ok': False, 'error': 'teacher not found'}), 404

        rs = t.get('roles') or []
        has_dept = any(str(r.get('department') or '').strip() == dept for r in rs)
        if has_dept:
            return jsonify({'ok': False, 'error': 'teacher already in this department'}), 400

        # 该部门的新 order 放到末尾
        max_order = 0
        for tt in teachers:
            for r in (tt.get('roles') or []):
                if str(r.get('department') or '').strip() == dept:
                    max_order = max(max_order, _safe_int(r.get('order'), default=0))

        new_role = {
            'department': dept,
            'position': '（待设置）',
            'order': max_order + 1,
        }
        t.setdefault('roles', []).append(new_role)
        write_json(DATA_TEACHERS, teachers)

        return jsonify({
            'ok': True,
            'role': {
                'roleKey': _role_key(t, new_role),
                'teacherId': t.get('id'),
                'name': t.get('name'),
                'department': dept,
                'position': new_role['position'],
                'order': new_role['order'],
            }
        })

    @app.post('/admin/teachers/departments/remove-role')
    @login_required
    def teachers_departments_remove_role():
        payload = request.get_json(silent=True) or {}
        role_key = str(payload.get('roleKey') or '').strip()
        if not role_key or '::' not in role_key:
            return jsonify({'ok': False, 'error': 'missing roleKey'}), 400

        try:
            teacher_id, dept, pos = role_key.split('::', 2)
        except Exception:
            return jsonify({'ok': False, 'error': 'invalid roleKey'}), 400

        teachers = load_json(DATA_TEACHERS)
        found = False
        removed = 0
        for t in teachers:
            if str(t.get('id') or '') != teacher_id:
                continue
            roles = t.get('roles') or []
            new_roles = []
            for r in roles:
                if str(r.get('department') or '') == dept and str(r.get('position') or '') == pos:
                    removed += 1
                    found = True
                    continue
                new_roles.append(r)
            t['roles'] = new_roles
            break

        if not found:
            return jsonify({'ok': False, 'error': 'role not found'}), 404

        write_json(DATA_TEACHERS, teachers)
        return jsonify({'ok': True, 'removed': removed})

    @app.post('/admin/teachers/apply-rules')
    @login_required
    def teachers_apply_rules():
        teachers = load_json(DATA_TEACHERS)
        teachers = _normalize_all_teachers(teachers)
        write_json(DATA_TEACHERS, teachers)
        flash('已对全库应用规则（理论组归类/岗位去重/管理部固定顺序），并写入 teachers.json（已备份）', 'ok')
        return redirect(url_for('teachers_list'))

    @app.get('/admin/teachers/<tid>')
    @login_required
    def teacher_edit(tid: str):
        teachers = load_json(DATA_TEACHERS)
        t = next((x for x in teachers if x.get('id') == tid), None)
        if not t:
            flash('未找到该教师', 'error')
            return redirect(url_for('teachers_list'))
        return render_template('teacher_edit.html', teacher=t)

    @app.post('/admin/teachers/<tid>')
    @login_required
    def teacher_edit_post(tid: str):
        teachers = load_json(DATA_TEACHERS)
        t = next((x for x in teachers if x.get('id') == tid), None)
        if not t:
            flash('未找到该教师', 'error')
            return redirect(url_for('teachers_list'))

        action = request.form.get('action') or 'save_all'

        # 通用字段
        t['photo'] = request.form.get('photo') or ''
        t['shortSummary'] = request.form.get('shortSummary') or ''
        t['bio'] = request.form.get('bio') or ''

        tm = _load_teachers_manage_module()
        apply_rules = (request.form.get('apply_rules') or 'on') in {'1', 'on', 'true', 'True', 'yes', 'YES'}
        apply_mgmt = (request.form.get('apply_mgmt_order') or 'on') in {'1', 'on', 'true', 'True', 'yes', 'YES'}

        if action == 'add_role':
            dept = getattr(tm, 'clean_dept')(request.form.get('new_department') or '')
            pos = getattr(tm, 'norm_line')(request.form.get('new_position') or '')
            order = _safe_int(request.form.get('new_order'), default=10**9)
            if not dept or not pos:
                flash('新增岗位失败：department/position 不能为空', 'error')
                return redirect(url_for('teacher_edit', tid=tid))

            t.setdefault('roles', [])
            t['roles'].append({'department': dept, 'position': pos, 'order': order})

        else:
            # save_all: 从表单整体回写 roles
            depts = request.form.getlist('role_department')
            poss = request.form.getlist('role_position')
            orders = request.form.getlist('role_order')
            delete_idx = {int(x) for x in request.form.getlist('role_delete') if str(x).isdigit()}

            roles = []
            for i in range(min(len(depts), len(poss), len(orders))):
                if i in delete_idx:
                    continue
                dept = getattr(tm, 'clean_dept')(depts[i] or '')
                pos = getattr(tm, 'norm_line')(poss[i] or '')
                order = _safe_int(orders[i], default=10**9)
                if not dept or not pos:
                    # 跳过空行，避免误写入脏数据
                    continue
                roles.append({'department': dept, 'position': pos, 'order': order})
            t['roles'] = roles

        # 应用业务规则
        if apply_rules:
            getattr(tm, 'normalize_teacher_roles')(t)

        if apply_mgmt:
            _enforce_mgmt_order([t])

        # 统一按 order 排序便于查看
        t['roles'] = sorted(t.get('roles') or [], key=lambda x: _safe_int(x.get('order')))

        if not (t.get('roles') or []):
            flash('保存失败：roles 不能为空（至少保留一个岗位）', 'error')
            return redirect(url_for('teacher_edit', tid=tid))

        write_json(DATA_TEACHERS, teachers)
        flash('已保存（roles 已更新）', 'ok')
        return redirect(url_for('teacher_edit', tid=tid))

    @app.get('/admin/students')
    @login_required
    def students_list():
        students = load_json(DATA_STUDENTS)
        students = sorted(students, key=lambda x: (-(int(x.get('year') or 0)), str(x.get('name') or '')))
        return render_template('students_list.html', students=students)

    @app.post('/admin/students/apply-rules')
    @login_required
    def students_apply_rules():
        students = load_json(DATA_STUDENTS)
        students, errs = _normalize_students(students)
        if errs:
            flash('校验未通过，未写入 students.json。错误示例：' + ('；'.join(errs[:6])), 'error')
            return redirect(url_for('students_list'))

        write_json(DATA_STUDENTS, students)
        flash('已对全库应用规则（字段规范化/清理 admissions/排序），并写入 students.json（已备份）', 'ok')
        return redirect(url_for('students_list'))

    @app.post('/admin/students/add')
    @login_required
    def students_add():
        sm = _load_students_manage_module()
        norm_line = getattr(sm, 'norm_line')

        name = norm_line(request.form.get('name') or '').replace(' ', '')
        school = norm_line(request.form.get('school') or '')
        major = norm_line(request.form.get('major') or '')
        year_raw = norm_line(request.form.get('year') or '')
        photo = norm_line(request.form.get('photo') or '')

        if not name or not school or not major:
            flash('新增失败：name/school/major 不能为空', 'error')
            return redirect(url_for('students_list'))

        year = None
        if year_raw:
            try:
                year = int(year_raw)
            except Exception:
                flash('新增失败：year 必须是整数', 'error')
                return redirect(url_for('students_list'))

        students = load_json(DATA_STUDENTS)
        if any((s.get('name') or '') == name for s in students if isinstance(s, dict)):
            flash('新增失败：已存在同名学生（请先到编辑页修改）', 'error')
            return redirect(url_for('students_list'))

        sid = getattr(sm, 'canonical_id')(name, school, year)
        existing_ids = {s.get('id') for s in students if isinstance(s, dict)}
        base = sid
        n = 1
        while sid in existing_ids:
            sid = f"{base}_{n}"
            n += 1

        entry = {
            'id': sid,
            'name': name,
            'school': school,
            'major': major,
            'year': year,
            'photo': photo,
            'admissions': [],
        }
        students.append(entry)
        students, errs = _normalize_students(students)
        if errs:
            flash('新增失败：数据校验未通过：' + ('；'.join(errs[:6])), 'error')
            return redirect(url_for('students_list'))

        write_json(DATA_STUDENTS, students)
        flash('已新增优秀考生', 'ok')
        return redirect(url_for('student_edit', sid=sid))

    @app.get('/admin/students/<sid>')
    @login_required
    def student_edit(sid: str):
        students = load_json(DATA_STUDENTS)
        s = next((x for x in students if x.get('id') == sid), None)
        if not s:
            flash('未找到该学生', 'error')
            return redirect(url_for('students_list'))
        return render_template('student_edit.html', student=s)

    @app.post('/admin/students/<sid>')
    @login_required
    def student_edit_post(sid: str):
        sm = _load_students_manage_module()
        norm_line = getattr(sm, 'norm_line')

        students = load_json(DATA_STUDENTS)
        s = next((x for x in students if x.get('id') == sid), None)
        if not s:
            flash('未找到该学生', 'error')
            return redirect(url_for('students_list'))

        action = request.form.get('action') or 'save_all'

        # 基本信息
        s['name'] = norm_line(request.form.get('name') or '').replace(' ', '')
        s['school'] = norm_line(request.form.get('school') or '')
        s['major'] = norm_line(request.form.get('major') or '')
        s['photo'] = norm_line(request.form.get('photo') or '')
        year_raw = norm_line(request.form.get('year') or '')
        if year_raw:
            try:
                s['year'] = int(year_raw)
            except Exception:
                flash('保存失败：year 必须是整数', 'error')
                return redirect(url_for('student_edit', sid=sid))
        else:
            s['year'] = None

        if not s.get('name') or not s.get('school') or not s.get('major'):
            flash('保存失败：name/school/major 不能为空', 'error')
            return redirect(url_for('student_edit', sid=sid))

        if action == 'add_admission':
            img = norm_line(request.form.get('new_image') or '')
            note = norm_line(request.form.get('new_note') or '')
            watermarked = (request.form.get('new_watermarked') or '') in {'1', 'on', 'true', 'True', 'yes', 'YES'}
            if not img:
                flash('新增录取截图失败：image 不能为空', 'error')
                return redirect(url_for('student_edit', sid=sid))
            s.setdefault('admissions', [])
            s['admissions'].append({'image': img, 'watermarked': bool(watermarked), 'note': note})
        else:
            # save_all: 回写 admissions
            images = request.form.getlist('admission_image')
            notes = request.form.getlist('admission_note')
            deletes = {int(x) for x in request.form.getlist('admission_delete') if str(x).isdigit()}
            watermarked_set = {int(x) for x in request.form.getlist('admission_watermarked') if str(x).isdigit()}

            admissions = []
            n_rows = min(len(images), len(notes))
            for i in range(n_rows):
                if i in deletes:
                    continue
                img = norm_line(images[i] or '')
                if not img:
                    continue
                admissions.append({
                    'image': img,
                    'watermarked': i in watermarked_set,
                    'note': norm_line(notes[i] or ''),
                })
            s['admissions'] = admissions

        students, errs = _normalize_students(students)
        if errs:
            flash('保存失败：数据校验未通过：' + ('；'.join(errs[:6])), 'error')
            return redirect(url_for('student_edit', sid=sid))

        write_json(DATA_STUDENTS, students)
        flash('已保存学生信息', 'ok')
        return redirect(url_for('student_edit', sid=sid))

    @app.post('/admin/students/<sid>/delete')
    @login_required
    def student_delete(sid: str):
        students = load_json(DATA_STUDENTS)
        before = len(students)
        students = [x for x in students if x.get('id') != sid]
        if len(students) == before:
            flash('未找到该学生', 'error')
            return redirect(url_for('students_list'))

        students, errs = _normalize_students(students)
        if errs:
            flash('删除后校验未通过，未写入：' + ('；'.join(errs[:6])), 'error')
            return redirect(url_for('students_list'))

        write_json(DATA_STUDENTS, students)
        flash('已删除学生', 'ok')
        return redirect(url_for('students_list'))

    return app


if __name__ == '__main__':
    app = create_app()
    host = os.environ.get('ADMIN_HOST', '127.0.0.1')
    port = int(os.environ.get('ADMIN_PORT', '5050'))
    debug = (os.environ.get('ADMIN_DEBUG') or '').strip() in {'1', 'true', 'True', 'yes', 'YES'}

    if (os.environ.get('ADMIN_SECRET_KEY') or '').strip() == '':
        print('WARNING: 未设置 ADMIN_SECRET_KEY，session 安全性不足。请设置一个随机长字符串。')
    if (os.environ.get('ADMIN_PASSWORD') or '').strip() == '':
        print('WARNING: 未设置 ADMIN_PASSWORD，将无法登录后台。')
    if host not in {'127.0.0.1', 'localhost'}:
        print(f'WARNING: 当前监听 {host}:{port}，请确保仅在可信内网使用并配置访问限制。')
    if debug:
        print('WARNING: ADMIN_DEBUG 已开启（Flask debug 模式）。请勿在对外环境启用。')

    app.run(host=host, port=port, debug=debug)
